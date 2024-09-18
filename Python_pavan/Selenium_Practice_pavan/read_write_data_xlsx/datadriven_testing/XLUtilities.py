import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(file,sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    row_count = sheet.max_row
    return row_count
def get_col_count(file,sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    column_count = sheet.max_column
    return column_count
def read_data(file,sheet,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    excel_data = sheet.cell(rownum,colnum).value
    return excel_data
def write_data(file,sheet,rownum,colnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    sheet.cell(rownum,colnum).value= data
    workbook.save()
def fill_green_color(file,sheet,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    greenfill = PatternFill(start_color ='60b212',end_color ='60b212',fill_type='solid')
    sheet.cell(rownum,colnum).fill = greenfill
    workbook.save(file)
def fill_red_color(file,sheet,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    redfill = PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(rownum,colnum).fill = redfill






